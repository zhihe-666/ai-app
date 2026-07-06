"""
迭代数据统计 API 路由

提供三个端点：
1. POST /api/stats/upload  — 上传 xlsx → 解析 → 返回统计数据
2. POST /api/stats/write-bitable — 将统计结果写入飞书多维表格
3. POST /api/stats/export — 导出统计结果 xlsx
"""

import logging
import os
import tempfile
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

from services.stats_engine import calculate_stats, parse_project_xlsx
from services.feishu_client import (
    get_bitable_records,
    batch_update_bitable,
    create_bitable_record,
    LarkCliError,
)

logger = logging.getLogger(__name__)

iteration_stats_bp = Blueprint("iteration_stats", __name__)

# 飞书多维表格配置
BASE_TOKEN = "B5exbr9CpafAW9sMEFkcydvvnRg"
TABLE_ID = "tbllFUFZyqhcUKZP"

# ── 固定项目名称→TL 映射表 ──
# 来源：飞书 wiki 表格，人工维护。不再依赖正则解析 bitable 中的版本号/日期
PROJECT_MAP = [
    ("DPP双周迭代", "樊少"),
    ("Dsearch搜索引擎版本迭代", "樊少"),
    ("Dgraph", "樊少"),
    ("交易搜索", "天央"),
    ("商业化", "天央"),
    ("社区搜索", "啊俊"),
    ("交易推荐", "三白"),
    ("用户&营销算法", "培成"),
    ("增长算法迭代", "培成"),
    ("95分算法", "岱锋"),
    ("国际算法迭代", "岱锋"),
    ("社区推荐", "则明"),
]
PROJECT_NAME_TO_TL = {name: tl for name, tl in PROJECT_MAP}


def _find_project_by_bitable_name(bitable_name_raw) -> str | None:
    """从 bitable 的项目名称（含版本号、链接格式）反向查找对应的标准项目名

    策略：去掉 markdown 链接格式 [xxx](url) → xxx，然后遍历 PROJECT_MAP
    看 bitable 名中是否包含标准项目名（子串匹配）
    """
    # 去掉 markdown 链接格式 [xxx](url) → xxx
    if isinstance(bitable_name_raw, str) and "[" in bitable_name_raw:
        name = bitable_name_raw.split("[")[-1].split("]")[0]
    else:
        name = str(bitable_name_raw) if bitable_name_raw else ""

    for std_name in PROJECT_NAME_TO_TL:
        if std_name in name:
            return std_name
    return None


def _fetch_bitable_project_map():
    """从飞书多维表格获取项目名称→TL 映射列表

    项目名和 TL 取自 PROJECT_MAP（硬编码），
    仅从 bitable 获取 record_id 用于写入更新。
    """
    try:
        records = get_bitable_records(BASE_TOKEN, TABLE_ID)
    except Exception as e:
        logger.warning(f"获取多维表格记录失败（降级处理）: {e}")
        return []

    projects = []
    for rec in records:
        name_raw = rec["fields"].get("项目名称", "")
        std_name = _find_project_by_bitable_name(name_raw)
        if std_name:
            projects.append({
                "record_id": rec["record_id"],
                "project_name": std_name,
                "tl": PROJECT_NAME_TO_TL[std_name],
            })

    logger.info(f"多维表格项目匹配结果: {[(p['project_name'], p['tl']) for p in projects]}")
    return projects


def _match_bitable_project(xlsx_name: str, bitable_projects: list) -> dict | None:
    """从 PROJECT_MAP 匹配 xlsx 中的项目名到标准项目名和 TL

    匹配 xlsx_name → PROJECT_MAP 标准名称（子串匹配）
    """
    if not bitable_projects:
        return None

    for std_name in PROJECT_NAME_TO_TL:
        if std_name in xlsx_name or xlsx_name in std_name:
            # 从 bitable_projects 找到对应 record（匹配 project_name）
            for bp in bitable_projects:
                if bp["project_name"] == std_name:
                    return bp
            # 没找到 record_id 也返回标准名+TL
            return {"project_name": std_name, "tl": PROJECT_NAME_TO_TL[std_name], "record_id": ""}

    return None


@iteration_stats_bp.route("/projects", methods=["GET"])
def list_projects():
    """获取多维表格中的现有项目列表（项目名称 + TL + record_id）

    项目名和 TL 取自 PROJECT_MAP（硬编码），
    仅从 bitable 获取 record_id。
    """
    try:
        records = get_bitable_records(BASE_TOKEN, TABLE_ID)
        projects = []
        for rec in records:
            name_raw = rec["fields"].get("项目名称", "")
            std_name = _find_project_by_bitable_name(name_raw)
            if std_name:
                projects.append({
                    "record_id": rec["record_id"],
                    "project_name": std_name,
                    "tl": PROJECT_NAME_TO_TL[std_name],
                })

        return jsonify({"success": True, "projects": projects})
    except LarkCliError as e:
        return jsonify({"error": f"飞书 API 错误: {e}"}), 502
    except Exception as e:
        logger.exception("获取项目列表失败")
        return jsonify({"error": f"服务器错误: {e}"}), 500


@iteration_stats_bp.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@iteration_stats_bp.route("/upload", methods=["POST"])
def upload_stats():
    """上传 xlsx 文件 → 解析统计 → 合并飞书项目TL信息 → 返回结果

    Request: multipart/form-data with file field "files" (one or more)
    Response: JSON with project_stats list and raw_stats per file
    """
    if "files" not in request.files:
        return jsonify({"error": "缺少 files 字段"}), 400

    uploaded_files = request.files.getlist("files")
    if not uploaded_files:
        return jsonify({"error": "未上传任何文件"}), 400

    # 1. 获取飞书多维表格中的项目列表（用于合并标准名称 + TL）
    bitable_projects = _fetch_bitable_project_map()
    logger.info(f"从多维表格获取到 {len(bitable_projects)} 个项目")

    results = []
    project_stats = []

    for f in uploaded_files:
        if not f.filename.endswith((".xlsx", ".xls")):
            continue

        # 保存到临时文件
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        try:
            f.save(tmp.name)
            tmp.close()

            df = parse_project_xlsx(tmp.name)
            raw = calculate_stats(df)

            filename = f.filename
            xlsx_name = raw.get("project_name") or filename.replace(".xlsx", "")

            # 从 bitable 项目列表中合并标准名称 + TL
            matched = _match_bitable_project(xlsx_name, bitable_projects)
            if matched:
                project_name = matched["project_name"]
                tl = matched["tl"]
            else:
                project_name = xlsx_name
                tl = ""

            row = {
                "project_name": project_name,
                "tl": tl,
                "source_file": filename,
                "total": raw["total_requirements"],
                "engineering": raw["engineering_requirements"],
                "aicoding": raw["aicoding_count"],
                "aicoding_ratio": raw["aicoding_ratio"],
                "sdd": raw["sdd_count"],
                "sdd_ratio": raw["sdd_ratio"],
                "e2e": raw["endtoend_count"],
                "detail_count": len(raw.get("details", [])),
            }
            results.append(row)
            project_stats.append(row)

        except Exception as e:
            logger.exception(f"处理 {f.filename} 失败")
            results.append({
                "source_file": f.filename,
                "error": str(e),
            })
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

    if not results:
        return jsonify({"error": "没有有效的 xlsx 文件"}), 400

    version = request.form.get("version", "")
    return jsonify({
        "success": True,
        "rows": project_stats,
        "version": version,
        "raw_results": results,
    })


@iteration_stats_bp.route("/write-bitable", methods=["POST"])
def write_bitable():
    """将客户端统计结果写入飞书多维表格

    Request JSON:
    {
        "project_stats": [
            {
                "project_name": "DPP双周迭代",
                "tl": "樊少(Fanshawe)",
                "total": 10,
                "engineering": 8,
                "aicoding": 5,
                "aicoding_ratio": 62.5,
                "sdd": 3,
                "sdd_ratio": 60.0,
                "e2e": 1
            },
            ...
        ]
    }

    Response: {"success": true, "updated": N, "errors": [...]}
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "缺少请求体"}), 400
    # 前端统一发 rows，兼容旧字段 project_stats
    project_stats = data.get("project_stats") or data.get("rows", [])
    if not project_stats:
        return jsonify({"error": "缺少 rows / project_stats 字段"}), 400

    try:
        # 1. 获取当前多维表格的所有记录 → 建立 std_name → record_id 映射
        records = get_bitable_records(BASE_TOKEN, TABLE_ID)
        logger.info(f"多维表格现有记录数: {len(records)}")

        # 用 PROJECT_MAP 中的标准名反向匹配 bitable 记录
        std_to_record_id = {}
        for rec in records:
            name_raw = rec["fields"].get("项目名称", "")
            std_name = _find_project_by_bitable_name(name_raw)
            if std_name:
                if std_name not in std_to_record_id:
                    std_to_record_id[std_name] = rec["record_id"]

        logger.info(f"标准名称→记录ID映射: {std_to_record_id}")

        # 2. 匹配：前端传入的 project_name → PROJECT_MAP 标准名
        records_data = []
        matched = 0
        unmatched = []
        used_ids = set()

        for ps in project_stats:
            pname = ps.get("project_name", "").strip()
            if not pname:
                continue

            record_id = None

            # 在 PROJECT_MAP 中找标准名（子串匹配）
            std_found = None
            for std_name in PROJECT_NAME_TO_TL:
                if std_name in pname or pname in std_name:
                    std_found = std_name
                    break

            if std_found and std_found in std_to_record_id:
                record_id = std_to_record_id[std_found]

            if record_id:
                if record_id not in used_ids:
                    # 用标准名和 TL 覆盖前端传入的值
                    ps["project_name"] = std_found
                    ps["tl"] = PROJECT_NAME_TO_TL[std_found]
                    records_data.append({
                        "record_id": record_id,
                        "stats": ps,
                    })
                    used_ids.add(record_id)
                    matched += 1
                else:
                    logger.warning(f"项目 '{pname}' 已匹配过记录 {record_id}，跳过")
            else:
                unmatched.append(pname)

        # 3c. 自动创建"合计"记录（bitable 中不存在时）
        summary_count = sum(1 for u in unmatched if u == "合计")
        if summary_count > 0:
            summary_row = next((ps for ps in project_stats if ps.get("project_name") == "合计"), None)
            if summary_row:
                try:
                    created = create_bitable_record(BASE_TOKEN, TABLE_ID, summary_row)
                    created_id = created.get("record_id", "")
                    if created_id:
                        records_data.append({
                            "record_id": created_id,
                            "stats": summary_row,
                        })
                        matched += 1
                        unmatched = [u for u in unmatched if u != "合计"]
                        logger.info(f"已自动创建合计行记录: {created_id}")
                except Exception as e:
                    logger.warning(f"自动创建合计行记录失败: {e}")

        if matched == 0:
            return jsonify({
                "success": False,
                "error": "没有匹配到任何项目记录。请确保项目名称与多维表格一致。",
                "unmatched": unmatched,
                "available_projects": sorted(PROJECT_NAME_TO_TL.keys()),
            }), 400

        # 4. 批量更新
        result = batch_update_bitable(BASE_TOKEN, TABLE_ID, records_data)

        return jsonify({
            "success": result["success"],
            "updated_count": result["updated"],
            "matched": matched,
            "unmatched": unmatched,
            "create_bitable_record": summary_count > 0,
            "errors": result["errors"],
        })

    except LarkCliError as e:
        return jsonify({"error": f"飞书 API 错误: {e}"}), 502
    except Exception as e:
        logger.exception("写入多维表格失败")
        return jsonify({"error": f"服务器错误: {e}"}), 500


@iteration_stats_bp.route("/export", methods=["POST"])
def export_stats():
    """导出统计结果为 xlsx 文件

    Request JSON:
    {
        "project_stats": [...]  (与 write-bitable 格式一致)
    }

    Response: xlsx file download
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "缺少请求体"}), 400
    project_stats = data.get("project_stats") or data.get("rows", [])
    if not project_stats:
        return jsonify({"error": "缺少 rows / project_stats 字段"}), 400

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "迭代数据统计"

        # 表头
        headers = [
            "项目名称", "TL", "总需求数【完全排期】", "算法工程需求",
            "AIcoding需求数", "AICoding需求占比", "SDD需求数",
            "SDD需求占比", "端到端需求数",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行
        data_font = Font(size=11)
        data_align = Alignment(horizontal="center", vertical="center")

        for row_idx, ps in enumerate(project_stats, 2):
            # aicoding_ratio 和 sdd_ratio 可能是带 "%" 的字符串（如 "25.0%"），直接使用
            a_ratio = ps.get("aicoding_ratio", "")
            s_ratio = ps.get("sdd_ratio", "")
            # 如果还是数值，格式化为百分比
            if isinstance(a_ratio, (int, float)):
                a_ratio = f"{a_ratio:.1f}%"
            if isinstance(s_ratio, (int, float)):
                s_ratio = f"{s_ratio:.1f}%"
            values = [
                ps.get("project_name", ""),
                ps.get("tl", ""),
                ps.get("total", 0),
                ps.get("engineering", 0),
                ps.get("aicoding", 0),
                a_ratio,
                ps.get("sdd", 0),
                s_ratio,
                ps.get("e2e", 0),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        # 列宽
        col_widths = [25, 20, 18, 14, 14, 16, 12, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 保存到临时文件
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.close()

        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=f"迭代数据统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.exception("导出 xlsx 失败")
        return jsonify({"error": f"导出失败: {e}"}), 500
